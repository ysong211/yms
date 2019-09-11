import xlrd
from openpyxl import load_workbook


# 打开表格
myWorkbook1 = xlrd.open_workbook('usa06.xls')
# 获取工作表list。
mySheets1 = myWorkbook1.sheets()
# 通过索引顺序获取
mySheet1 = mySheets1[0]
# 获取行数
nrows1 = mySheet1.nrows


# 打开表格
myWorkbook2 = xlrd.open_workbook('usa11.xls')
# 获取工作表list。
mySheets2 = myWorkbook2.sheets()
# 通过索引顺序获取
mySheet2 = mySheets2[0]
# 获取行数
nrows2 = mySheet2.nrows

# 生成一个已存在的wookbook对象
wb = load_workbook("usa06.xls")

for i in range(0, nrows1-1):
    Value = mySheet2.cell_value(i, 3)
    wb1 = wb.active  # 激活sheet
    wb1.cell(nrows1-1, 3, Value)


wb.save("usa06.xls")





# print(nrows)
# # 生成一个已存在的wookbook对象
# wb = load_workbook("usa11.xls")
# # 直接获取单元格数据，行数 列数，行数和列数都是从0开始计数。
# for i in range(0, nrows-1):
#     myCellValue = mySheet.cell_value(i, 3)
#     wb1 = wb.active  # 激活sheet
#
#     wb1.cell(nrows-1, 3, myCellValue)  # 往sheet中的第二行第二列写入‘pass2’的数据
#
#
# wb.save("10月关键词.xlsx")
