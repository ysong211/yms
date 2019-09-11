import xlrd
from openpyxl import load_workbook
import xlwt


# 创建工作workbook
workbook = xlwt.Workbook(encoding="utf-8")

# 创建工作表worksheet,填入表名
worksheet = workbook.add_sheet('sheet1')

a = 0
names = ['usa10.xls', 'usa12.xls',  'usa13.xls',  'usa14.xls']
for name in names:
    # 打开表格
    myWorkbook = xlrd.open_workbook(name)
    # 获取工作表list。
    mySheets = myWorkbook.sheets()
    # 通过索引顺序获取
    mySheet = mySheets[0]
    # 获取行数
    nrows = mySheet.nrows
    print(nrows)

    for i in range(0, nrows):
        a += 1
        # 直接获取单元格数据，行数 列数，行数和列数都是从0开始计数。
        Value = mySheet.cell_value(i, 3)
        worksheet.write(a, 1, 1)
        worksheet.write(a, 2, 2)
        worksheet.write(a, 3, Value)
print(a)
workbook.save('usa8_{}.xls'.format(a))



# # 打开表格
# myWorkbook2 = xlrd.open_workbook('usa11.xls')
# # 获取工作表list。
# mySheets2 = myWorkbook2.sheets()
# # 通过索引顺序获取
# mySheet2 = mySheets2[0]
# # 获取行数
# nrows2 = mySheet2.nrows
#
# # 生成一个已存在的wookbook对象
# wb = load_workbook("usa06.xls")
#
# for i in range(0, nrows1-1):
#     Value = mySheet2.cell_value(i, 3)
#     wb1 = wb.active  # 激活sheet
#     # wb1.cell(nrows1-1, 3, Value)
#     wb1.write(nrows1-1, 3, Value)
#
#
# wb.save("usa06.xls")





# print(nrows)
# # 生成一个已存在的wookbook对象
# wb = load_workbook("usa11.xls")
# # 直接获取单元格数据，行数 列数，行数和列数都是从0开始计数。
# for i in range(0, nrows-1):
#     myCellValue = mySheet.cell_value(i, 3)
#     wb1 = wb.active  # 激活sheet
#
#     wb1.cell(nrows-1, 3, myCellValue)  # 往sheet中的第二行第二列写入‘pass2’的数据
#
#
# wb.save("10月关键词.xlsx")
